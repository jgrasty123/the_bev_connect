#!/usr/bin/env python3
"""
Build assets/data/weship-zones.json from the WeShip zone map for our
Ventura (93003) origin.

    python3 tools/build_weship_zones.py <weship-zone-map.csv|.xlsx>

WHICH SOURCE
------------
Use the zone map that WeShip supplies with the 2025 rate program — most likely
a tab in `The_BroBasket_Proposed_pricing_grids_2025_v2_-_weship.xlsx`, the same
workbook the rate grid was extracted from.

Do NOT substitute a UPS or FedEx zone chart. WeShip assigns its own zones. If
their boundaries differ from a carrier's, a substituted chart mis-assigns zones
*silently* — the zone-to-rate checks still pass (zone 8 at 30 lb is still
$35.19), so the error never surfaces in testing. It only shows up as invoices
that do not reconcile against quotes.

EXPECTED INPUT
--------------
Two columns are needed: a destination ZIP prefix (or a hyphenated range) and a
ground zone. Column names are detected loosely; anything containing "zip",
"dest", or "prefix" is treated as the key, and anything containing "zone" or
"ground" as the value. Ranges like `010-041` are expanded to one entry each.

Rows with no ground service (`-`, blank, `[1]`) are skipped and reported —
those prefixes then fall through to "request a quote" on the page, which is the
correct behaviour.

If the real file does not fit this shape, adjust `detect_columns()` rather than
reshaping the source by hand.
"""

import csv
import json
import os
import re
import sys

OUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "assets", "data", "weship-zones.json",
)

# Zones present in weship-ground-rates-2025.json
VALID_ZONES = {2, 3, 4, 5, 6, 7, 8, 44, 45, 46}

PREFIX_RE = re.compile(r"^\s*(\d{3})(?:\s*-\s*(\d{3}))?\s*$")
NO_SERVICE = {"-", "", "[1]", "n/a", "na"}


def load_rows(path):
    """Yield lists of cell values from a CSV or XLSX source."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl needed for xlsx: pip install openpyxl --break-system-packages")
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                yield [("" if c is None else str(c)) for c in row]
    else:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            for row in csv.reader(fh):
                yield row


def detect_columns(rows):
    """Find (key_idx, zone_idx) from a header row, else fall back to 0, 1."""
    for row in rows[:40]:
        low = [c.strip().lower() for c in row]
        key = next((i for i, c in enumerate(low)
                    if any(t in c for t in ("zip", "dest", "prefix"))), None)
        zone = next((i for i, c in enumerate(low)
                     if i != key and any(t in c for t in ("zone", "ground"))), None)
        if key is not None and zone is not None:
            return key, zone
    return 0, 1


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)

    src = sys.argv[1]
    if not os.path.exists(src):
        sys.exit("Source file not found: %s" % src)

    rows = list(load_rows(src))
    key_i, zone_i = detect_columns(rows)
    print("Using column %d as ZIP prefix, column %d as zone." % (key_i, zone_i))

    zones, skipped, unexpected = {}, [], []

    for row in rows:
        if len(row) <= max(key_i, zone_i):
            continue
        m = PREFIX_RE.match(str(row[key_i]).strip().strip('"'))
        if not m:
            continue

        raw = str(row[zone_i]).strip().strip('"')
        if raw.lower() in NO_SERVICE:
            skipped.append(row[key_i].strip())
            continue
        try:
            zone = int(float(raw))
        except ValueError:
            unexpected.append((row[key_i].strip(), raw))
            continue
        if zone not in VALID_ZONES:
            unexpected.append((row[key_i].strip(), raw))
            continue

        start = int(m.group(1))
        end = int(m.group(2)) if m.group(2) else start
        for n in range(start, end + 1):
            zones["%03d" % n] = zone

    if not zones:
        sys.exit(
            "No zone rows parsed from %s. Check the column detection above, or "
            "adjust detect_columns() to match the real layout." % src
        )

    payload = {
        "_meta": {
            "origin": "Ventura, CA 93003",
            "program": "WeShip Express 2025",
            "service": "Ground Residential",
            "source_file": os.path.basename(src),
            "prefix_count": len(zones),
            "note": "WeShip zone map. Do not substitute a UPS or FedEx zone chart.",
        },
        "zones": zones,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1, sort_keys=True)
        fh.write("\n")

    print("Wrote %s" % OUT_PATH)
    print("  prefixes mapped : %d" % len(zones))
    print("  distinct zones  : %s" % sorted(set(zones.values())))
    if skipped:
        print("  no ground svc   : %d rows (%s...)"
              % (len(skipped), ", ".join(skipped[:6])))
    if unexpected:
        print("  UNEXPECTED values (review before shipping):")
        for pfx, val in unexpected[:20]:
            print("    %s -> %r" % (pfx, val))

    local = zones.get("930")
    if local is not None and local > 3:
        print("  WARNING: prefix 930 resolved to zone %d — expected 2 for a "
              "Ventura origin. Wrong origin map?" % local)


if __name__ == "__main__":
    main()
